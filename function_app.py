import azure.functions as func
import logging
import io
import pandas as pd
from openpyxl import load_workbook
import re
import zipfile
from openai import AzureOpenAI
from urllib.parse import quote
from pathlib import Path
import os
from dotenv import load_dotenv
import boto3
from botocore.config import Config
import json
import time

# .envファイルから環境変数を読み込む
load_dotenv()

# FunctionAppの初期化
app = func.FunctionApp(http_auth_level=func.AuthLevel.ANONYMOUS)

# --- LLMサービス設定 ---
llm_service = os.getenv("LLM_SERVICE", "AWS")

# --- Azure OpenAI Service 接続情報 ---
azure_api_key = os.getenv("AZURE_OPENAI_API_KEY")
azure_endpoint = os.getenv("AZURE_OPENAI_ENDPOINT")
azure_api_version = os.getenv("AZURE_OPENAI_API_VERSION")
azure_deployment = os.getenv("AZURE_OPENAI_DEPLOYMENT")

# --- AWS Bedrock 接続情報 ---
aws_region = os.getenv("AWS_REGION")
aws_access_key_id = os.getenv("AWS_ACCESS_KEY_ID")
aws_secret_access_key = os.getenv("AWS_SECRET_ACCESS_KEY")
aws_bedrock_model_id = os.getenv("AWS_BEDROCK_MODEL_ID")

# LLMクライアントの初期化用変数
azure_client = None
bedrock_client = None

# 必須環境変数のチェック関数
def validate_env():
    if llm_service == "AZURE":
        # Azure OpenAIに必要な環境変数がすべて設定されているか確認
        required = [azure_api_key, azure_endpoint, azure_api_version, azure_deployment]
        if not all(required):
            raise ValueError("Azure OpenAI の必須環境変数が設定されていません。")
    elif llm_service == "AWS":
        # AWS Bedrockに必要な環境変数がすべて設定されているか確認
        required = [aws_region, aws_access_key_id, aws_secret_access_key, aws_bedrock_model_id]
        if not all(required):
            raise ValueError("AWS Bedrock の必須環境変数が設定されていません。")
    else:
        # サポートされていないLLMサービスが指定された場合のエラー
        raise ValueError(f"無効なLLMサービスが指定されました: {llm_service}")

# LLMクライアントの初期化関数
def initialize_client():
    global azure_client, bedrock_client
    validate_env()  # 環境変数の妥当性をチェック

    if llm_service == "AZURE":
        # Azure OpenAIクライアントの初期化
        azure_client = AzureOpenAI(
            api_version=azure_api_version,
            azure_endpoint=azure_endpoint,
            api_key=azure_api_key,
        )
    elif llm_service == "AWS":
        # AWS Bedrockクライアントの初期化（タイムアウト設定付き）
        config = Config(read_timeout=600, connect_timeout=60)
        bedrock_client = boto3.client(
            "bedrock-runtime",
            region_name=aws_region,
            aws_access_key_id=aws_access_key_id,
            aws_secret_access_key=aws_secret_access_key,
            config=config,
        )

# LLMサービスを呼び出す共通関数
def call_llm(system_prompt: str, user_prompt: str, max_retries: int = 5) -> str:
    """
    指定されたLLMサービス（AzureまたはAWS）を使ってプロンプトを送信し、応答を取得する。
    system_prompt: システムプロンプト（モデルの振る舞いを定義）
    user_prompt: ユーザーからの入力
    max_retries: 最大リトライ回数
    戻り値: モデルからの応答テキスト
    """
    global azure_client, bedrock_client
    
    # クライアントが未初期化の場合は初期化する
    if llm_service == "AZURE" and azure_client is None:
        initialize_client()
    elif llm_service == "AWS" and bedrock_client is None:
        initialize_client()
    
    for attempt in range(max_retries):
        try:
            if llm_service == "AZURE":
                # Azure OpenAIにチャット形式でリクエストを送信
                response = azure_client.chat.completions.create(
                    model=azure_deployment,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt},
                    ],
                    max_completion_tokens=32768,
                )
                return response.choices[0].message.content

            elif llm_service == "AWS":
                # AWS BedrockにConverse APIでリクエストを送信
                response = bedrock_client.converse(
                    modelId=aws_bedrock_model_id,
                    messages=[{"role": "user", "content": [{"text": user_prompt}]}],
                    system=[{"text": system_prompt}],
                    inferenceConfig={"maxTokens": 64000},
                )
                # レスポンスの構造を確認してから取得
                if 'output' in response and 'message' in response['output']:
                    return response['output']['message']['content'][0]['text']
                else:
                    logging.error(f"予期しないレスポンス構造: {json.dumps(response, ensure_ascii=False)}")
                    raise RuntimeError("AWS Bedrockからの応答形式が不正です。")

        except Exception as e:
            error_message = str(e)
            # ThrottlingExceptionの場合はリトライ
            if "ThrottlingException" in error_message or "Too many requests" in error_message:
                if attempt < max_retries - 1:
                    wait_time = (2 ** attempt) + (attempt * 2)  # エクスポネンシャルバックオフ
                    logging.warning(f"{llm_service} API レート制限エラー。{wait_time}秒後にリトライします（{attempt + 1}/{max_retries}）")
                    time.sleep(wait_time)
                    continue
                else:
                    logging.error(f"{llm_service} API呼び出しが最大リトライ回数に達しました")
                    raise RuntimeError(f"{llm_service} APIのレート制限エラー。しばらく待ってから再試行してください。")
            else:
                # その他のエラーは即座に失敗
                logging.error(f"{llm_service} API呼び出し中にエラーが発生しました: {error_message}")
                raise RuntimeError(f"{llm_service} API呼び出しに失敗しました: {error_message}")
    
    raise RuntimeError(f"{llm_service} API呼び出しに失敗しました")


def structuring(prompt: str) -> str:
    system_prompt = '''
        あなたは業務システムの設計書を解析し、構造化されたMarkdownドキュメントを作成する専門家です。

        【タスク】
        提供されたExcelシートの生データから、設計情報を抽出し、読みやすく整理してください。
        データは表形式、箇条書き、自由記述など様々な形式で提供されます。

        【出力要件】
        - データの内容に応じて最適な形式で構造化（表、リスト、セクション分けなど）
        - 表形式のデータは可能な限りMarkdown表として出力
        - 処理フローや機能仕様がある場合は、以下を抽出：
          - 処理ID/番号、処理名/機能名
          - トリガー/実行条件、処理内容/動作
          - 使用するデータ/テーブル/API、遷移先/出力
        - 画面項目定義がある場合は、項目ID/番号、項目名、型、必須/任意、初期値、制約条件などを整理
        - その他の情報も内容に応じて適切に構造化

        【重要】識別情報の保持
        - 章番号、セクション番号、処理番号、項目番号、項目IDなどの識別情報は必ず保持
        - 見出しには番号を含める（例：「## 2.1 ユーザー登録画面」「### 処理1: 初期表示」）
        - 表の列に番号やIDがある場合は必ず含める
        - これらの識別情報は後続のテスト設計でトレーサビリティに使用されます

        【記述ルール】
        - 意味のない行・空欄・重複情報は除外
        - 箇条書きは `-` を使用（`<br>`タグは使用しない）
        - ヘッダー行や列名は適切に認識して活用
        - 複数の表が含まれる場合は、見出しで区切る
        - 出力形式はMarkdown
    '''
    return call_llm(system_prompt, prompt)

def extract_test_perspectives(prompt: str) -> str:
    system_prompt = '''
        あなたはソフトウェアテストの専門家です。提供された設計書からテスト観点を抽出してください。

        【タスク】
        設計書の内容を分析し、機能・処理単位でテスト観点を整理してください。

        【出力形式】
        - 機能・処理単位で `##` セクションに分けてください
        - 各機能・処理について以下を記述：
          - **仕様概要**：機能の目的、入出力、制約条件
          - **業務ルール**：業務上の制約、分岐条件、依存関係
          - **テスト観点**：確認すべきポイント（正常系、異常系、境界値、初期値、エラー処理など）
          - **レビューポイント**：この機能で特に注意すべき点、見落としやすいリスク
          - **要確認事項**：設計書に記載がなく、テスト設計に必要な情報（境界値、エラー処理、前提条件など）
        
        【記述ルール】
        - 個別の画面項目ごとではなく、機能・処理単位でまとめて記述
        - 入力チェックや制約条件は要約して記載
        - モードや状態による分岐は明示
        - レビューポイントは「なぜこのテストが重要か」を簡潔に説明
        - 要確認事項は「設計書に明記されていないが、テストに必要な情報」を質問形式で記載
        - 要確認事項がない場合は「なし」と記載
        - 出力形式はMarkdown
    '''
    return call_llm(system_prompt, prompt)

def create_test_spec(prompt: str) -> str:
    system_prompt = '''
        あなたはソフトウェア品質保証の専門家です。
        提供された設計書とテスト観点をもとに、実務レベルのテスト仕様書を作成してください。

        【情報の使い分け】
        - 設計書：テストケースの具体的な内容、項目順序、詳細仕様を参照
        - テスト観点：レビューポイントを参照

        【最重要】設計書の記載順序を厳守：
        - 設計書に項目定義表がある場合：上から順に各項目をテストケース化
        - 処理フローがある場合：処理番号順にテストケース化
        - 各項目は「初期値→正常系→境界値→異常系」の順で展開
        - 初期値・デフォルト値がある場合は必ず確認テストを作成

        【出力形式】
        以下の6列構成のMarkdown表で出力：
        - No: 連番（1から開始）
        - 大区分: 機能名や画面名（例：「ユーザー登録画面」「データ取込処理」）
        - 中区分: 処理名や項目名（例：「初期表示」「氏名入力」「メール送信」）
        - テストケース: 具体的なテスト内容
        - 期待結果: 確認事項（1行1結果）
        - トレース元: 設計書の参照箇所（章番号、項目ID、処理番号など）

        【記述ルール】
        - テストケース列：「～を確認する」「～をテストする」で統一
        - 期待結果列：「～であること」「～されること」で統一
        - 大区分列：機能名や画面名を記載（数値IDではなく名称）
        - 中区分列：処理名や項目名を記載（数値IDではなく名称）
        - トレース元列：設計書の章番号、項目ID、処理番号などを明記
        - 項目名は設計書の正式名称を使用
        - 同じ大区分・中区分が連続する場合は該当欄を空白にして省略可
        - 表形式のみ出力（説明文は不要）

        【禁止事項】
        - 曖昧な表現（「その他の項目」「適切に」など）
        - 複数確認事項の1行化
        - 設計書の順序を無視した並び替え
        - 語尾の不統一
    '''
    return call_llm(system_prompt, prompt)

def structuring_screen_list(prompt: str) -> str:
    system_prompt = '''
        あなたは業務システムの画面一覧を解析し、構造化されたMarkdownドキュメントを作成する専門家です。

        【タスク】
        提供されたExcelシートから画面一覧情報を抽出し、読みやすく整理してください。

        【出力要件】
        - 画面ID、画面名、画面種別、URL、概要などを抽出
        - Markdown表形式で出力
        - 画面IDは必ず保持（後続処理で画面遷移定義とマッチングに使用）
        - 画面の分類・グループがあれば見出しで整理

        【記述ルール】
        - 意味のない行・空欄は除外
        - ヘッダー行を適切に認識
        - 出力形式はMarkdown
    '''
    return call_llm(system_prompt, prompt)

def structuring_transition(prompt: str) -> str:
    system_prompt = '''
        あなたは業務システムの画面遷移定義を解析し、構造化されたMarkdownドキュメントを作成する専門家です。

        【タスク】
        提供されたExcelシートから画面遷移情報を抽出し、読みやすく整理してください。

        【出力要件】
        - 遷移元画面ID/名、遷移先画面ID/名、遷移条件、トリガー（ボタン名など）を抽出
        - Markdown表形式で出力
        - 画面IDは必ず保持（画面一覧とのマッチングに使用）
        - 業務フローごとに見出しで整理

        【記述ルール】
        - 意味のない行・空欄は除外
        - ヘッダー行を適切に認識
        - 出力形式はMarkdown
    '''
    return call_llm(system_prompt, prompt)

def create_integration_test_spec(prompt: str) -> str:
    system_prompt = '''
        あなたは結合テストの専門家です。構造化詳細設計書と画面関連情報から、実行可能な結合テスト仕様書を作成してください。

        【重要】スコープの理解
        - 構造化詳細設計書に記載されている画面のみを対象とする
        - 画面一覧・画面遷移図には全システムの情報が含まれているが、詳細設計書に記載がある画面のみを使用
        - 手順：
          1. まず構造化詳細設計書から対象画面を特定する
          2. 画面遷移図から、特定した画面間の遷移のみを抽出する
          3. 設計書に記載のない画面は、遷移図にあってもテストシナリオに含めない
        - 例：設計書に「ログイン画面」「商品検索画面」のみある場合、遷移図に「管理画面」があってもそれは使用しない

        【タスク】
        設計書の処理フローと画面遷移情報から、画面間をまたぐ業務シナリオを抽出し、具体的なテストデータを含むテストケースに展開してください。
        
        【テストデータの生成】
        - 入力項目には具体的な値を記述（例：ユーザーID="admin"、パスワード="Pass123"）
        - 正常値、境界値、異常値のパターンを具体例で示す
        - 設計書に制約があればそれに合わせた値を生成（例：桁数が10桁なら"1234567890"）
        - 日付は"2024-01-15"形式、金額は"10000"など具体的に

        【出力形式】
        以下の列構成のMarkdown表で出力：
        - テストNo: テストケースID（A-01, A-02など）
        - 画面1: 開始画面名
        - 機能/操作/状況1: 画面1での具体的操作（テストデータ含む）
        - 画面2: 遷移先画面名
        - 機能/操作/状況2: 画面2での具体的操作（テストデータ含む）
        - 画面3: 遷移先画面名
        - 機能/操作/状況3: 画面3での具体的操作（テストデータ含む）
        - 画面4: 遷移先画面名
        - 機能/操作/状況4: 画面4での具体的操作（テストデータ含む）
        - 確認内容: 具体的な確認事項

        【列数の統一】
        - 全てのテストケースで上記の列構成を統一すること
        - 画面遷移が少ない場合は該当列を空欄にする
        - 例：2画面のみの場合、画面3～4と機能/操作/状況3～4は空欄
        - 4画面を超える場合は複数のテストケースに分割する

        【テストケース生成の観点】
        - 正常系：代表的な業務フローを具体的なデータでテスト
        - 境界値：桁数上限、下限、最大値、最小値など
        - 異常系：入力エラー、権限エラー、データ不存在など

        【記述ルール】
        - 操作内容は「ボタン名 + 具体的な入力値」を明記
        - 例：「ユーザーID欄に"admin"、パスワード欄に"Pass123!"を入力し、「ログイン」ボタンをクリック」
        - 確認内容は「～であること」「～されること」で統一
        - 表形式のみ出力（説明文は不要）
        
        【禁止事項】
        - 構造化詳細設計書に記載のない画面をテストシナリオに含めること
        - 画面遷移図のみに存在する画面を使用すること
    '''
    return call_llm(system_prompt, prompt)

@app.route(route="upload", methods=["POST"])
def upload(req: func.HttpRequest) -> func.HttpResponse:
    test_type = req.form.get("testType", "unit")
    
    if test_type == "unit":
        return generate_unit_test(req)
    elif test_type == "integration":
        return generate_integration_test(req)
    else:
        return func.HttpResponse("無効なテストタイプです", status_code=400)

def generate_unit_test(req: func.HttpRequest) -> func.HttpResponse:
    try:
        file = req.files.get("documentFile")
        if not file:
            return func.HttpResponse("ファイルがアップロードされていません", status_code=400)
        
        file_bytes = file.read()
        filename = file.filename
        
        if not filename.endswith('.xlsx'):
            return func.HttpResponse("Excelファイル(.xlsx)のみ対応しています", status_code=400)
            
    except Exception as e:
        logging.error(f"ファイル取得エラー: {e}")
        return func.HttpResponse("ファイルの取得に失敗しました", status_code=400)

    logging.info(f"{filename} を受信しました。単体テスト生成を開始します。")

    try:
        # アップロードされたExcelファイル（バイナリ）をメモリ上で読み込み、全シートを辞書形式で取得
        # すべてのシートが {シート名: DataFrame} の形式で格納される
        # header=Noneで全行をデータとして読み込み、_clean_sheet()でヘッダーを設定
        excel_data = pd.read_excel(io.BytesIO(file_bytes), sheet_name=None, header=None)

        # Markdown構造化のためのリスト初期化
        toc_list = [] # 目次(Table of Contents)用のリスト
        md_sheets = [] # 各シートのmd文字列を格納するリスト

        # --- 各シートの処理 ---
        for sheet_name, df in excel_data.items():
            # 目次用のアンカーを生成 (GitHub-flavored)
            anchor = re.sub(r'[^a-z0-9-]', '', sheet_name.strip().lower().replace(' ', '-'))
            toc_list.append(f'- [{sheet_name}](#{anchor})')
            
            sheet_content = f"## {sheet_name}\n\n"

            # --- すべてのシートをAIで構造化 ---
            logging.info(f"「{sheet_name}」シートをAIで構造化します。")
            try:
                # DataFrameを行ごとにテキスト化（セル区切りを明示）
                raw_text = '\n'.join(df.apply(lambda row: ' | '.join(row.astype(str).fillna('')), axis=1))
                
                structuring_prompt = f'''
                    --- Excelシート「{sheet_name}」 ---
                    {raw_text}
                '''
                structured_content = structuring(structuring_prompt)
                sheet_content += structured_content
                
            except Exception as e:
                logging.error(f"AIによるシート構造化中にエラー: {e}")
                sheet_content += "（AIによる構造化に失敗しました）"
            
            md_sheets.append(sheet_content)

        # --- 1. 全体を結合して最終的なMarkdown設計書を生成 ---
        logging.info("全シートの処理が完了。最終的な設計書を組み立てます。")
        md_output_first = f"# {filename}\n\n"
        md_output_first += "## 目次\n\n"
        md_output_first += "\n".join(toc_list)
        md_output_first += "\n\n---\n\n"
        md_output_first += "\n\n---\n\n".join(md_sheets)
        logging.info("Markdown設計書をメモリ上に生成しました。")
        
        # --- 2. AIによるテスト観点抽出 ---
        logging.info("設計書全体をAIに渡し、テスト観点を抽出します。")
        extract_test_perspectives_prompt = f'''
            --- 設計書 ---
            {md_output_first}
        '''
        md_output_second = extract_test_perspectives(extract_test_perspectives_prompt)
        logging.info("テスト観点抽出が完了し、メモリ上に保持しました。")

        # --- 3. AIによるテスト仕様書生成 ---
        logging.info("設計書全体をAIに渡し、テスト仕様書を生成します。")
        test_gen_prompt = f'''
            --- 設計書 ---
            {md_output_first}
            
            --- テスト観点 ---
            {md_output_second}
        '''

        md_output_third = create_test_spec(test_gen_prompt)
        logging.info("テスト仕様書の生成が完了し、メモリ上に保持しました。")

        # --- 4. テスト仕様書(Markdown)をExcelに変換 ---
        logging.info("テスト仕様書をMarkdownからExcel形式に変換します。")
        
        # Markdownから表部分だけを抽出
        md_lines = [line.strip() for line in md_output_third.splitlines() if line.strip().startswith("|")]
        
        if not md_lines:
            logging.error("テスト仕様書にMarkdown表が見つかりませんでした")
            return func.HttpResponse("テスト仕様書の生成に失敗しました（表形式が見つかりません）", status_code=500)
        
        tsv_text = "\n".join([line.strip("|").replace("|", "\t") for line in md_lines])

        # DataFrame化
        df = pd.read_csv(io.StringIO(tsv_text), sep="\t")
        df.columns = [col.strip() for col in df.columns]
        
        # 必須列の存在確認（6列構成）
        required_columns = ["No", "大区分", "中区分", "テストケース", "期待結果", "トレース元"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            logging.error(f"必須列が不足しています: {missing_columns}")
            return func.HttpResponse(f"テスト仕様書の形式が不正です（不足列: {', '.join(missing_columns)}）", status_code=500)
        
        # 既存テンプレートを読み込み
        template_path = "単体テスト仕様書.xlsx"
        wb = load_workbook(template_path)
        ws = wb.active

        # マッピング定義（DataFrame列名 → Excel列番号）
        column_map = {
            "No": 1,           # A列
            "大区分": 2,       # B列
            "中区分": 6,       # F列
            "テストケース": 10, # J列
            "期待結果": 23,     # W列
            "トレース元": 42    # AP列
        }

        # DataFrameをA11,B11,F11,J11,W11,AP11に書き込み
        start_row = 11
        for i, row in enumerate(df.itertuples(index=False), start=start_row):
            for col_name, excel_col in column_map.items():
                if col_name in df.columns:
                    ws.cell(row=i, column=excel_col, value=getattr(row, col_name))

        # バッファに保存（メモリ上）
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        excel_bytes = excel_buffer.read()
        
        logging.info("テンプレートExcelへの書き込みが完了しました。")

        # --- 5. 全成果物をZIPファイルにまとめる ---
        logging.info("全成果物をZIPファイルにまとめています。")
        base_name = Path(filename).stem
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr(f"{base_name}_構造化設計書.md", md_output_first.encode('utf-8'))
            zip_file.writestr(f"{base_name}_テスト観点.md", md_output_second.encode('utf-8'))
            zip_file.writestr(f"{base_name}_テスト仕様書.md", md_output_third.encode('utf-8'))
            zip_file.writestr(f"{base_name}_テスト仕様書.xlsx", excel_bytes)
        
        zip_buffer.seek(0)
        zip_bytes = zip_buffer.read()
        logging.info("ZIPファイルの作成が完了しました。")
        
        # --- 6. ユーザーへの返却（ZIPファイル） ---
        output_filename = f"テスト仕様書_{Path(filename).stem}.zip"
        encoded_filename = quote(output_filename)
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Type": "application/zip",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
        return func.HttpResponse(zip_bytes, status_code=200, headers=headers)

    except ValueError as ve:
        logging.error(f"設定エラー: {ve}")
        return func.HttpResponse(str(ve), status_code=500)
    except Exception as e:
        logging.error(f"処理全体で予期せぬエラーが発生: {e}")
        return func.HttpResponse("処理中にサーバーエラーが発生しました", status_code=500)

def generate_integration_test(req: func.HttpRequest) -> func.HttpResponse:
    try:
        structured_design_files = req.files.getlist("structuredDesignFiles")
        transition_diagram_file = req.files.get("transitionDiagramFile")
        
        if not structured_design_files or not transition_diagram_file:
            return func.HttpResponse("必須ファイルが不足しています", status_code=400)
            
    except Exception as e:
        logging.error(f"ファイル取得エラー: {e}")
        return func.HttpResponse("ファイルの取得に失敗しました", status_code=400)

    logging.info("結合テスト生成を開始します。")

    try:
        # 複数の構造化詳細設計書を読み込み、結合
        structured_design_md = ""
        for design_file in structured_design_files:
            content = design_file.read().decode('utf-8')
            structured_design_md += f"\n\n# {design_file.filename}\n\n{content}\n\n---\n\n"
        logging.info(f"{len(structured_design_files)}件の構造化詳細設計書を読み込みました。")
        
        # 画面一覧/画面遷移図を読み込み、AIで構造化
        logging.info("画面一覧/画面遷移図（Excel）をAIで構造化します。")
        transition_data = pd.read_excel(io.BytesIO(transition_diagram_file.read()), sheet_name=None, header=None)
        transition_md = ""
        for sheet_name, df in transition_data.items():
            raw_text = '\n'.join(df.apply(lambda row: ' | '.join(row.astype(str).fillna('')), axis=1))
            prompt = f'--- 画面一覧/画面遷移図「{sheet_name}」 ---\n{raw_text}'
            transition_md += f"## {sheet_name}\n\n{structuring_transition(prompt)}\n\n"
        logging.info("画面一覧/画面遷移図の構造化が完了しました。")
        
        # 結合テスト仕様書を直接生成
        logging.info("結合テスト仕様書を生成します。")
        test_spec_prompt = f'''
            --- 構造化詳細設計書 ---
            {structured_design_md}
            
            --- 画面一覧/画面遷移図 ---
            {transition_md}
        '''
        test_spec_md = create_integration_test_spec(test_spec_prompt)
        logging.info("結合テスト仕様書の生成が完了しました。")
        
        # ZIPファイルにまとめる
        logging.info("全成果物をZIPファイルにまとめています。")
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr("1_画面関連情報.md", transition_md.encode('utf-8'))
            zip_file.writestr("2_結合テスト仕様書.md", test_spec_md.encode('utf-8'))
        
        zip_buffer.seek(0)
        zip_bytes = zip_buffer.read()
        logging.info("ZIPファイルの作成が完了しました。")
        
        output_filename = "結合テスト仕様書.zip"
        encoded_filename = quote(output_filename)
        headers = {
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
            "Content-Type": "application/zip",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
        return func.HttpResponse(zip_bytes, status_code=200, headers=headers)

    except ValueError as ve:
        logging.error(f"設定エラー: {ve}")
        return func.HttpResponse(str(ve), status_code=500)
    except Exception as e:
        logging.error(f"処理全体で予期せぬエラーが発生: {e}")
        return func.HttpResponse("処理中にサーバーエラーが発生しました", status_code=500)